from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
GREEN_DARK = "1B5E20"
GREEN_MED = "2E7D32"
WHITE = "FFFFFF"
GRAY_LIGHT = "F5F5F5"
GRAY_MED = "E0E0E0"
RED = "C62828"

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
title_font = Font(name="Calibri", bold=True, size=16, color=GREEN_DARK)
subtitle_font = Font(name="Calibri", size=10, color="666666")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
money_font = Font(name="Calibri", size=10, color="1B5E20")
thin_border = Border(
    left=Side(style="thin", color=GRAY_MED),
    right=Side(style="thin", color=GRAY_MED),
    top=Side(style="thin", color=GRAY_MED),
    bottom=Side(style="thin", color=GRAY_MED),
)

def hdr(ws, row, cols, fill_color=GREEN_DARK):
    f = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def row_style(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if alt:
            cell.fill = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")

def total_row(ws, row, cols, fill_color=GREEN_MED):
    f = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        cell.fill = f
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

def section_header(ws, row, cols, text, fill_color):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, cols+1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

def write_data(ws, start_row, data, cols, money_cols=None):
    money_cols = money_cols or []
    for i, d in enumerate(data):
        r = start_row + i
        for j, val in enumerate(d):
            cell = ws.cell(row=r, column=j+1, value=val)
            if j in money_cols and isinstance(val, (int, float)):
                cell.number_format = '$#,##0'
                cell.font = money_font
        row_style(ws, r, cols, alt=(i % 2 == 1))
    return start_row + len(data)

def write_total(ws, row, cols, data, money_cols, fill_color=GREEN_MED, label="TOTAL"):
    ws.cell(row=row, column=1, value=label)
    for j in money_cols:
        ws.cell(row=row, column=j+1, value=sum(d[j] for d in data if isinstance(d[j], (int, float)))).number_format = '$#,##0'
    total_row(ws, row, cols, fill_color)
    return row + 1

def auto_width(ws, cols, min_width=12, max_width=45):
    for c in range(1, cols+1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: MASTER COMPARISON
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "MASTER COMPARISON"
ws.sheet_properties.tabColor = GREEN_DARK

r = 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=18)
ws.cell(row=r, column=1, value="GINA'S TENNIS WORLD — DEPLOYMENT OPTIONS MASTER COMPARISON").font = Font(name="Calibri", bold=True, size=18, color=GREEN_DARK)
ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=18)
ws.cell(row=r, column=1, value="Last Updated: June 2026 | 16 deployment options compared across 20+ categories").font = subtitle_font
ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 2

# Cost comparison section
section_header(ws, r, 18, "MONTHLY COST COMPARISON (Low / Typical / High)", "1565C0")
r += 1
cost_headers = ["Option", "Monthly Low", "Monthly Typical", "Monthly High", "Annual Typical", "Setup Time", "Cold Starts", "Auto-Scaling", "DB Backups", "Email", "Monitoring", "SSL/HTTPS", "GDPR Ready", "Dev Experience", "Vendor Lock-in", "Keep FastAPI", "Python Support", "Uptime SLA"]
for i, h in enumerate(cost_headers, 1):
    ws.cell(row=r, column=i, value=h)
hdr(ws, r, 18, "1565C0")
r += 1

cost_data = [
    ["Current (Vercel+Render Free)", 14, 14, 14, 168, "0 (done)", "30-60 sec", "No", "Manual", "Gmail (limited)", "None", "Auto (Vercel)", "Partial", "Good", "Medium", "Yes", "Yes", "99.9%"],
    ["Full AWS", 40, 65, 99, 780, "2-3 days", "None", "Yes", "7-day auto", "SES (pro)", "CloudWatch", "ACM (auto)", "Yes", "Complex", "High", "Yes", "Yes", "99.99%"],
    ["Hybrid (Vercel+AWS)", 49, 68, 88, 816, "1-2 days", "None (Vercel)", "Yes", "7-day auto", "SES (pro)", "Mixed", "Auto", "Yes", "Good", "Medium", "Yes", "Yes", "99.99%"],
    ["Vercel+Render (Upgraded)", 52, 80, 120, 960, "1-2 hours", "None", "Limited", "Daily auto", "Resend", "Sentry", "Auto", "Partial", "Great", "Low", "Yes", "Yes", "99.9%"],
    ["Vercel+Railway", 30, 50, 98, 600, "2-4 hours", "None", "Yes", "Auto", "Resend", "Sentry", "Auto", "Partial", "Great", "Low", "Yes", "Yes", "99.9%"],
    ["DigitalOcean", 20, 35, 57, 420, "2-4 hours", "None", "Limited", "Auto", "Resend", "Built-in", "Auto", "Partial", "Good", "Low", "Yes", "Yes", "99.9%"],
    ["Google Cloud (GCP)", 17, 40, 77, 480, "1-2 days", "Slight", "Yes", "Auto", "SendGrid", "Cloud Mon", "Managed", "Yes", "Complex", "Medium", "Yes", "Yes", "99.95%"],
    ["Microsoft Azure", 27, 40, 60, 480, "1-2 days", "None", "Yes", "Auto", "SendGrid", "Azure Mon", "Managed", "Yes", "Complex", "Medium", "Yes", "Yes", "99.9%"],
    ["Fly.io", 6, 25, 65, 300, "2-4 hours", "None", "Yes", "Auto", "Resend", "Sentry", "Auto", "Partial", "Good", "Low", "Yes", "Yes (Docker)", "99.9%"],
    ["Heroku", 16, 30, 71, 360, "1-2 hours", "30 sec (Eco)", "Limited", "Daily", "SendGrid", "Basic", "Auto", "Partial", "Great", "Medium", "Yes", "Yes", "99.9%"],
    ["Netlify+Supabase", 44, 70, 110, 840, "3-5 days", "None", "Yes", "Daily auto", "Resend", "Sentry", "Auto", "Partial", "Great", "High", "No (Deno)", "No (Deno)", "99.9%"],
    ["Vercel+Supabase", 45, 75, 111, 900, "3-5 days", "None", "Yes", "Daily auto", "Resend", "Sentry", "Auto", "Partial", "Great", "High", "No (Deno)", "No (Deno)", "99.9%"],
    ["Cloudflare+Neon", 24, 35, 56, 420, "2-3 days", "None", "Yes", "Auto", "Resend", "CF Analytics", "Auto", "Yes", "Good", "Medium", "No (JS)", "No (JS)", "99.9%"],
    ["Self-Hosted VPS", 6, 10, 26, 120, "1-2 days", "None", "No", "Manual", "Resend", "Uptime Kuma", "Certbot", "Manual", "Manual", "None", "Yes", "Yes (Docker)", "None"],
    ["Vercel+PlanetScale", 74, 95, 122, 1140, "2-3 days", "None", "Yes", "Auto", "Resend", "Sentry", "Auto", "Partial", "Great", "Medium", "Yes (MySQL)", "Yes", "99.9%"],
    ["Coolify (Self-Host)", 10, 15, 30, 180, "2-4 hours", "None", "No", "Auto", "Resend", "Built-in", "Auto (Coolify)", "Partial", "Great", "Low", "Yes", "Yes (Docker)", "99.9%"],
    ["Vercel+Neon DB", 64, 85, 112, 1020, "2-3 days", "None", "Yes", "Auto", "Resend", "Sentry", "Auto", "Partial", "Good", "Medium", "Yes", "Yes", "99.9%"],
    ["Supabase Only", 45, 70, 111, 840, "3-5 days", "None", "Yes", "Daily auto", "Resend", "Sentry", "Auto", "Partial", "Great", "High", "No (Deno)", "No (Deno)", "99.9%"],
]

for i, d in enumerate(cost_data):
    r2 = r + i
    for j, val in enumerate(d):
        ws.cell(row=r2, column=j+1, value=val)
        if j in [1, 2, 3, 4] and isinstance(val, (int, float)):
            ws.cell(row=r2, column=j+1).number_format = '$#,##0'
    row_style(ws, r2, 18, alt=(i % 2 == 1))
    # Highlight current setup row
    if "Current" in str(d[0]):
        for c in range(1, 19):
            ws.cell(row=r2, column=c).font = Font(name="Calibri", bold=True, size=10, color=RED)

r = r + len(cost_data) + 2

# Feature matrix
section_header(ws, r, 18, "FEATURE MATRIX (✓ = Included | ✗ = Not included | ~ = Partial | $ = Paid add-on)", "6A1B9A")
r += 1
feat_headers = ["Feature", "Current", "Full AWS", "Hybrid\n(V+AWS)", "V+Render\nPro", "V+Railway", "DO", "GCP", "Azure", "Fly.io", "Heroku", "Net+Supa", "V+Supa", "CF+Neon", "VPS", "V+PScale", "Coolify", "V+Neon"]
for i, h in enumerate(feat_headers, 1):
    ws.cell(row=r, column=i, value=h)
hdr(ws, r, 18, "6A1B9A")
r += 1

features = [
    ["Auto-Scaling", "✗", "✓", "✓", "~", "✓", "~", "✓", "✓", "✓", "~", "✓", "✓", "✓", "✗", "~", "✓", "✓"],
    ["No Cold Starts", "✗", "✓", "✓", "✓", "✓", "✓", "~", "✓", "✓", "✗", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["DB Auto-Backups", "✗", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "✓", "✓", "✓"],
    ["Production Email", "✗", "✓", "✓", "$", "$", "$", "$", "$", "$", "$", "$", "$", "$", "$", "$", "$", "$"],
    ["Monitoring/Alerts", "✗", "✓", "✓", "$", "$", "✓", "✓", "✓", "$", "~", "$", "$", "✓", "$", "$", "✓", "$"],
    ["Custom Domain+SSL", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["Global CDN", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"],
    ["Preview Deploys", "✓", "✗", "✓", "✓", "✓", "✗", "✗", "✗", "✗", "✓", "✓", "✓", "✓", "✗", "✓", "✓", "✓"],
    ["Keep FastAPI", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "✗", "✗", "✓", "✓", "✓", "✓"],
    ["GDPR Compliance", "~", "✓", "✓", "~", "~", "~", "✓", "✓", "~", "~", "~", "~", "✓", "✗", "~", "~", "~"],
    ["DDoS Protection", "~", "✓", "✓", "~", "~", "~", "✓", "✓", "✓", "~", "✓", "✓", "✓", "✓", "~", "~", "~"],
    ["Managed Auth", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✓", "✓", "✗", "✗", "✗", "✗", "✗"],
    ["Real-time DB", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✓", "✓", "✗", "✗", "✗", "✗", "✗"],
    ["DB Branching", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✓", "✓", "✓", "✗", "✓", "✓", "✓"],
    ["Serverless DB", "✗", "✗", "✗", "✗", "✗", "✗", "✓", "✗", "✗", "✗", "✓", "✓", "✓", "✗", "✓", "✗", "✓"],
    ["Object Storage", "✗", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "✓", "✓", "✓", "✗", "✓", "✓", "✓"],
    ["Edge Functions", "✗", "✓", "✗", "✗", "✗", "✓", "✓", "✓", "✓", "✗", "✓", "✓", "✓", "✗", "✗", "✗", "✗"],
    ["99.9%+ SLA", "~", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "✓", "✓", "✓"],
    ["Easy Rollback", "✓", "~", "✓", "✓", "✓", "~", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "✓", "✓", "✓"],
    ["Team Collaboration", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "✓", "✓", "✓"],
]

for i, d in enumerate(features):
    r2 = r + i
    for j, val in enumerate(d):
        cell = ws.cell(row=r2, column=j+1, value=val)
    row_style(ws, r2, 18, alt=(i % 2 == 1))
    ws.cell(row=r2, column=1).font = bold_font
    for j in range(1, 18):
        cell = ws.cell(row=r2, column=j+1)
        v = cell.value
        if v == "✓":
            cell.font = Font(name="Calibri", size=10, color=GREEN_MED)
        elif v == "✗":
            cell.font = Font(name="Calibri", size=10, color=RED)
        elif v == "~":
            cell.font = Font(name="Calibri", size=10, color="FF6F00")
        elif v == "$":
            cell.font = Font(name="Calibri", size=10, color="1565C0")

auto_width(ws, 18, min_width=10, max_width=22)
ws.column_dimensions["A"].width = 22


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: CURRENT SETUP
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Current Setup")
ws2.sheet_properties.tabColor = "1565C0"

r = 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws2.cell(row=r, column=1, value="CURRENT SETUP: Vercel Free + Render Free + SQLite").font = title_font
ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 1
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws2.cell(row=r, column=1, value="What you're running right now | Total: ~$14/mo").font = subtitle_font
ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 2

for i, h in enumerate(["Category", "Service", "Plan", "Monthly", "Annual", "Notes"], 1):
    ws2.cell(row=r, column=i, value=h)
hdr(ws2, r, 6)
r += 1

data = [
    ["Frontend Hosting", "Vercel", "Hobby (Free)", 0, 0, "100GB bandwidth, serverless functions"],
    ["Backend Hosting", "Render", "Starter ($7/mo)", 7, 84, "1 vCPU, 512MB RAM, sleeps after 15min"],
    ["Database", "Render PostgreSQL", "Starter ($7/mo)", 7, 84, "256MB RAM, 1GB storage, sleeps"],
    ["Domain & SSL", "Vercel", "Free", 0, 0, "Automatic SSL, custom domain"],
    ["Email", "Gmail / SMTP", "Free", 0, 0, "Limited: 500 emails/day, spam risk"],
    ["File Storage", "Render Disk", "Starter", 0, 0, "Included with Render plan"],
    ["CDN", "Vercel Edge", "Free", 0, 0, "Included with Vercel"],
    ["Monitoring", "Console logs", "Free", 0, 0, "No dedicated monitoring"],
    ["Secrets", "Render Env Vars", "Free", 0, 0, "Environment variables in dashboard"],
    ["Payment Processing", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
]
r = write_data(ws2, r, data, 6, money_cols=[3, 4])
r = write_total(ws2, r, 6, data, [3, 4], label="TOTAL (est.)")
ws2.cell(row=r-1, column=6, value="Excludes Stripe fees")

r += 1
section_header(ws2, r, 6, "KNOWN LIMITATIONS", RED)
r += 1
lim_data = [
    ["", "Render sleeps after 15min inactivity", "Cold starts: 30-60 seconds", "Bad UX", "", ""],
    ["", "SQLite not suitable for production", "No concurrent writes", "Data corruption risk", "", ""],
    ["", "No automated backups", "Manual only", "Data loss risk", "", ""],
    ["", "Gmail SMTP: 500 emails/day", "No deliverability guarantees", "Emails go to spam", "", ""],
    ["", "No monitoring/alerting", "No error tracking", "Issues go unnoticed", "", ""],
    ["", "Vercel hobby: 100GB bandwidth", "Could hit limit with traffic", "Site goes down", "", ""],
]
for i, d in enumerate(lim_data):
    for j, val in enumerate(d):
        ws2.cell(row=r+i, column=j+1, value=val)
    row_style(ws2, r+i, 6, alt=(i % 2 == 1))
    ws2.cell(row=r+i, column=2).font = Font(name="Calibri", bold=True, size=10, color=RED)

auto_width(ws2, 6)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: ALL OPTIONS DETAIL
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("All Options Detail")
ws3.sheet_properties.tabColor = "E65100"

r = 1
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws3.cell(row=r, column=1, value="ALL 16 DEPLOYMENT OPTIONS — DETAILED COST BREAKDOWN").font = title_font
ws3.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 2

options = [
    ("Full AWS", "E65100", [
        ["Frontend", "AWS Amplify", "SSR Hosting", 0, 15, "Free tier: 5GB storage, 15GB transfer"],
        ["Backend", "AWS App Runner", "0.5 vCPU / 1 GB", 25, 35, "Auto-scaling, no cold starts"],
        ["Database", "AWS RDS PostgreSQL", "db.t4g.micro, 20GB", 12, 30, "Free tier eligible (12 months)"],
        ["Email", "AWS SES", "Production", 0, 5, "First 3,000 emails/mo free"],
        ["Domain & SSL", "Route 53 + ACM", "Custom domain", 1.50, 1.50, "ACM free, Route 53 ~$1.50"],
        ["File Storage", "AWS S3", "5 GB", 0.25, 2, "Profile photos, uploads"],
        ["CDN", "CloudFront", "Global", 1, 5, "Caches static assets"],
        ["Monitoring", "CloudWatch", "Logs + 3 alarms", 0, 5, "Application & DB monitoring"],
        ["Secrets", "AWS Secrets Manager", "4 secrets", 0.40, 0.40, "$0.10/secret/month"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Hybrid (Vercel+AWS)", "7B1FA2", [
        ["Frontend", "Vercel", "Pro ($20/mo)", 20, 20, "Unlimited bandwidth, preview deploys"],
        ["Backend", "AWS App Runner", "0.25 vCPU / 0.5 GB", 16, 25, "Auto-scaling, no cold starts"],
        ["Database", "AWS RDS PostgreSQL", "db.t4g.micro, 20GB", 12, 30, "Free tier eligible (12 months)"],
        ["Email", "AWS SES", "Production", 0, 5, "First 3,000 emails/mo free"],
        ["Domain & SSL", "Vercel + Route 53", "Custom domain", 0.50, 0.50, "Vercel free SSL + Route 53"],
        ["File Storage", "AWS S3", "5 GB", 0.25, 2, "Profile photos, uploads"],
        ["CDN", "Vercel Edge", "Included", 0, 0, "Vercel handles frontend CDN"],
        ["Monitoring", "Vercel + CloudWatch", "Basic", 0, 5, "Vercel analytics + CW for backend"],
        ["Secrets", "Vercel Env + AWS Secrets", "Mixed", 0.20, 0.20, "Vercel env vars + 2 AWS secrets"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Vercel+Render Pro", "00695C", [
        ["Frontend", "Vercel", "Pro ($20/mo)", 20, 20, "Unlimited bandwidth, analytics"],
        ["Backend", "Render", "Standard ($25/mo)", 25, 25, "2 GB RAM, no auto-sleep"],
        ["Database", "Render PostgreSQL", "Starter→Standard", 7, 20, "Upgrade when needed"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Vercel", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Render Disk", "Starter", 0.25, 2, "Persistent disk"],
        ["CDN", "Vercel Edge", "Included", 0, 0, "Global CDN"],
        ["Monitoring", "Sentry", "Free / Team ($26)", 0, 26, "Error tracking + performance"],
        ["Secrets", "Render Env Vars", "Free", 0, 0, "Encrypted env vars"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Vercel+Railway", "F57C00", [
        ["Frontend", "Vercel", "Pro ($20/mo)", 20, 20, "Unlimited bandwidth"],
        ["Backend", "Railway", "Usage-based", 5, 20, "Pay only for what you use"],
        ["Database", "Railway PostgreSQL", "Usage-based", 5, 10, "Managed Postgres, auto-backups"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Vercel", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Railway Volume", "Usage-based", 0.50, 2, "Persistent storage"],
        ["CDN", "Vercel Edge", "Included", 0, 0, "Global CDN"],
        ["Monitoring", "Railway + Sentry", "Free / Team", 0, 26, "Sentry free: 5K events/mo"],
        ["Secrets", "Railway Env Vars", "Free", 0, 0, "Encrypted env vars"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("DigitalOcean", "00838F", [
        ["Frontend", "DO App Platform", "Static (Free)", 0, 0, "3 free static sites"],
        ["Backend", "DO App Platform", "Basic ($5/mo)", 5, 20, "512MB-2GB RAM"],
        ["Database", "DO Managed DB", "Basic ($15/mo)", 15, 15, "1GB RAM, 10GB, auto-backups"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "DO + Let's Encrypt", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "DO Spaces", "5GB", 0.25, 2, "S3-compatible"],
        ["CDN", "DO Spaces CDN", "Included", 0, 0, "Free CDN with Spaces"],
        ["Monitoring", "DO Monitoring", "Free", 0, 0, "Built-in monitoring"],
        ["Secrets", "DO App Env Vars", "Free", 0, 0, "Encrypted env vars"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Google Cloud (GCP)", "4285F4", [
        ["Frontend", "Firebase Hosting", "Spark (Free)", 0, 0, "1GB storage, 10GB transfer/mo"],
        ["Backend", "Cloud Run", "1 vCPU / 512MB", 5, 20, "Pay-per-use, scales to zero"],
        ["Database", "Cloud SQL (Postgres)", "db-f1-micro, 10GB", 10, 30, "Shared vCPU, auto-backups"],
        ["Email", "SendGrid", "Free / Pro ($15)", 0, 15, "100 emails/day free"],
        ["Domain & SSL", "Cloud Domains + SSL", "Custom domain", 1, 1, "Managed SSL free"],
        ["File Storage", "Cloud Storage", "5 GB", 0.10, 1, "Profile photos, uploads"],
        ["CDN", "Cloud CDN", "Global", 1, 5, "Caches static assets"],
        ["Monitoring", "Cloud Monitoring", "Free tier", 0, 5, "1 MiB logs free"],
        ["Secrets", "Secret Manager", "6 secrets free", 0, 0.50, "6 secrets free"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Microsoft Azure", "0078D4", [
        ["Frontend", "Azure Static Web Apps", "Free", 0, 0, "100GB bandwidth, custom domains"],
        ["Backend", "Azure App Service", "Basic B1 ($13/mo)", 13, 13, "1 vCPU, 1.75GB RAM"],
        ["Database", "Azure DB for PostgreSQL", "Flexible Burstable", 13, 25, "1 vCore, 20GB, auto-backups"],
        ["Email", "SendGrid", "Free / Pro ($15)", 0, 15, "100 emails/day free"],
        ["Domain & SSL", "Azure DNS + SSL", "Custom domain", 0.50, 0.50, "Free managed SSL"],
        ["File Storage", "Azure Blob Storage", "5 GB", 0.10, 1, "Hot tier, LRS"],
        ["CDN", "Azure Front Door", "Global", 1, 5, "Microsoft's global CDN"],
        ["Monitoring", "Azure Monitor", "Basic", 0, 5, "5GB logs/mo free"],
        ["Secrets", "Azure Key Vault", "10 ops/mo free", 0, 1, "Secrets, keys, certificates"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Fly.io", "7B2D8E", [
        ["Frontend", "Fly.io (Next.js)", "Shared-cpu-1x 256MB", 1.94, 5, "Deploy as container"],
        ["Backend", "Fly.io (FastAPI)", "Shared-cpu-1x 256MB", 1.94, 5, "Deploy as container"],
        ["Database", "Fly.io Postgres", "Shared-cpu-1x 256MB", 1.94, 7, "Managed Postgres"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Fly.io", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Fly.io Volumes", "3 GB", 0.30, 2, "Persistent volumes"],
        ["CDN", "Fly.io Edge", "Included", 0, 0, "Global edge network"],
        ["Monitoring", "Fly.io + Sentry", "Free / Team", 0, 26, "Sentry free: 5K events/mo"],
        ["Secrets", "Fly.io Secrets", "Free", 0, 0, "Encrypted secrets"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Heroku", "430098", [
        ["Frontend", "Heroku (Next.js)", "Eco ($5/mo)", 5, 5, "Eco dyno, sleeps after inactivity"],
        ["Backend", "Heroku (FastAPI)", "Eco ($5/mo)", 5, 5, "Eco dyno, sleeps after inactivity"],
        ["Database", "Heroku Postgres", "Mini→Essential", 5, 15, "10K rows (Mini), upgrade to Essential"],
        ["Email", "SendGrid", "Free / Pro ($15)", 0, 15, "100 emails/day free"],
        ["Domain & SSL", "Heroku", "Free (Eco+)", 0, 0, "Automatic SSL on paid dynos"],
        ["File Storage", "S3 (external)", "External ($1-5)", 1, 5, "Need external S3 bucket"],
        ["CDN", "Heroku (CloudFront)", "Included", 0, 0, "Built-in CDN"],
        ["Monitoring", "Heroku + Sentry", "Free / Team", 0, 26, "Basic metrics included"],
        ["Secrets", "Heroku Config Vars", "Free", 0, 0, "Encrypted env vars"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Netlify+Supabase", "00C853", [
        ["Frontend", "Netlify", "Pro ($19/mo)", 19, 19, "1TB bandwidth, 100K builds/mo"],
        ["Backend", "Supabase Edge Functions", "Free / Pro ($25)", 0, 25, "Deno-based edge functions"],
        ["Database", "Supabase PostgreSQL", "Pro ($25/mo)", 25, 25, "8GB storage, auto-backups"],
        ["Auth", "Supabase Auth", "Included", 0, 0, "JWT, OAuth, MFA"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Netlify", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Supabase Storage", "Included (8GB)", 0, 0, "S3-compatible, CDN-backed"],
        ["CDN", "Netlify + Supabase CDN", "Included", 0, 0, "Global CDN"],
        ["Monitoring", "Netlify + Sentry", "Free / Team", 0, 26, "Netlify analytics in Pro"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Vercel+Supabase", "FF6D00", [
        ["Frontend", "Vercel", "Pro ($20/mo)", 20, 20, "Unlimited bandwidth"],
        ["Backend", "Supabase Edge Functions", "Free / Pro ($25)", 0, 25, "Deno-based edge functions"],
        ["Database", "Supabase PostgreSQL", "Pro ($25/mo)", 25, 25, "8GB storage, auto-backups"],
        ["Auth", "Supabase Auth", "Included", 0, 0, "JWT, OAuth, MFA"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Vercel", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Supabase Storage", "Included (8GB)", 0, 0, "S3-compatible, CDN-backed"],
        ["CDN", "Vercel + Supabase CDN", "Included", 0, 0, "Global CDN"],
        ["Monitoring", "Vercel + Sentry", "Free / Team", 0, 26, "Vercel analytics in Pro"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Cloudflare+Neon", "F57F17", [
        ["Frontend", "Cloudflare Pages", "Free / Pro ($20)", 0, 20, "500 builds/mo, unlimited bandwidth"],
        ["Backend", "Cloudflare Workers", "Paid ($5/mo)", 5, 10, "10M requests/mo included"],
        ["Database", "Neon PostgreSQL", "Pro ($19/mo)", 19, 19, "Serverless Postgres, branching"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Cloudflare", "Free", 0, 0, "Free SSL, DNS, proxy"],
        ["File Storage", "Cloudflare R2", "5 GB free", 0, 1, "S3-compatible, no egress fees"],
        ["CDN", "Cloudflare CDN", "Free", 0, 0, "300+ edge locations"],
        ["Monitoring", "Cloudflare Analytics", "Free", 0, 0, "Built-in analytics"],
        ["Secrets", "CF Workers Secrets", "Free", 0, 0, "Encrypted env vars"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Self-Hosted VPS", "BF360C", [
        ["Frontend", "Nginx on VPS", "Hetzner CX22 ($6/mo)", 6, 6, "2 vCPU, 4GB RAM, 40GB SSD"],
        ["Backend", "Same VPS (Docker)", "Included", 0, 0, "FastAPI in Docker"],
        ["Database", "PostgreSQL on VPS", "Included", 0, 0, "Self-managed on same server"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Let's Encrypt (Certbot)", "Free", 0, 0, "Free SSL, auto-renewal"],
        ["File Storage", "VPS Disk", "Included (40GB)", 0, 0, "SSD storage included"],
        ["CDN", "Cloudflare Free", "Free", 0, 0, "Free CDN + DDoS protection"],
        ["Monitoring", "Uptime Kuma (self-hosted)", "Free", 0, 0, "Self-hosted monitoring"],
        ["Secrets", "Docker Env Vars", "Free", 0, 0, "Encrypted in .env file"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Vercel+PlanetScale", "1A237E", [
        ["Frontend", "Vercel", "Pro ($20/mo)", 20, 20, "Unlimited bandwidth"],
        ["Backend", "Render", "Standard ($25/mo)", 25, 25, "2 GB RAM, no auto-sleep"],
        ["Database", "PlanetScale", "Scaler ($29/mo)", 29, 29, "5GB storage, 1B rows read/mo"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Vercel", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Render Disk", "Starter", 0.25, 2, "Persistent disk"],
        ["CDN", "Vercel Edge", "Included", 0, 0, "Global CDN"],
        ["Monitoring", "Sentry", "Free / Team ($26)", 0, 26, "Error tracking"],
        ["Secrets", "Vercel + Render Env", "Free", 0, 0, "Encrypted env vars"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Coolify (Self-Host)", "00695C", [
        ["VPS Server", "Hetzner CX32", "2 vCPU, 8GB RAM", 10, 10, "Enough for everything"],
        ["Frontend", "Coolify (Next.js)", "On VPS", 0, 0, "Deploy via Git, auto SSL"],
        ["Backend", "Coolify (FastAPI)", "On VPS", 0, 0, "Deploy via Git, Docker"],
        ["Database", "Coolify (PostgreSQL)", "On VPS", 0, 0, "One-click install, auto-backups"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Coolify + Let's Encrypt", "Free", 0, 0, "Automatic SSL, reverse proxy"],
        ["File Storage", "VPS Disk (80GB)", "Included", 0, 0, "SSD storage included"],
        ["CDN", "Cloudflare Free", "Free", 0, 0, "Free CDN + DDoS protection"],
        ["Monitoring", "Coolify Built-in", "Free", 0, 0, "Built-in monitoring"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Vercel+Neon DB", "4A148C", [
        ["Frontend", "Vercel", "Pro ($20/mo)", 20, 20, "Unlimited bandwidth"],
        ["Backend", "Render", "Standard ($25/mo)", 25, 25, "2 GB RAM, no auto-sleep"],
        ["Database", "Neon PostgreSQL", "Pro ($19/mo)", 19, 19, "Serverless Postgres, branching"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Vercel", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Render Disk", "Starter", 0.25, 2, "Persistent disk"],
        ["CDN", "Vercel Edge", "Included", 0, 0, "Global CDN"],
        ["Monitoring", "Sentry", "Free / Team ($26)", 0, 26, "Error tracking"],
        ["Secrets", "Vercel + Render Env", "Free", 0, 0, "Encrypted env vars"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
    ("Supabase Only", "33691E", [
        ["Frontend", "Vercel", "Pro ($20/mo)", 20, 20, "Still need Vercel for Next.js SSR"],
        ["Backend", "Supabase Edge Functions", "Pro ($25/mo)", 25, 25, "Deno-based edge functions"],
        ["Database", "Supabase PostgreSQL", "Pro ($25/mo)", 25, 25, "8GB storage, auto-backups"],
        ["Auth", "Supabase Auth", "Included in Pro", 0, 0, "JWT, OAuth, MFA, RLS"],
        ["Email", "Resend", "Free / Pro ($20)", 0, 20, "3,000 emails/mo free"],
        ["Domain & SSL", "Vercel", "Free", 0, 0, "Automatic SSL"],
        ["File Storage", "Supabase Storage", "Included (8GB)", 0, 0, "S3-compatible, CDN-backed"],
        ["CDN", "Vercel + Supabase CDN", "Included", 0, 0, "Global CDN"],
        ["Monitoring", "Supabase + Sentry", "Free / Team", 0, 26, "Supabase logs included"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction"],
    ]),
]

for opt_name, opt_color, opt_data in options:
    section_header(ws3, r, 8, f"OPTION: {opt_name}", opt_color)
    r += 1
    for i, h in enumerate(["Category", "Service", "Config", "Monthly Low", "Monthly High", "Annual Low", "Annual High", "Notes"], 1):
        ws3.cell(row=r, column=i, value=h)
    hdr(ws3, r, 8, opt_color)
    r += 1
    for i, d in enumerate(opt_data):
        for j, val in enumerate(d):
            ws3.cell(row=r, column=j+1, value=val)
            if j in [3, 4] and isinstance(val, (int, float)):
                ws3.cell(row=r, column=j+1).number_format = '$#,##0'
                ws3.cell(row=r, column=j+1).font = money_font
        # Add annual columns
        ws3.cell(row=r, column=6, value=d[3]*12 if isinstance(d[3], (int, float)) else d[3]).number_format = '$#,##0'
        ws3.cell(row=r, column=7, value=d[4]*12 if isinstance(d[4], (int, float)) else d[4]).number_format = '$#,##0'
        row_style(ws3, r, 8, alt=(i % 2 == 1))
        r += 1
    # Total row
    ws3.cell(row=r, column=1, value="TOTAL")
    ws3.cell(row=r, column=4, value=sum(d[3] for d in opt_data if isinstance(d[3], (int, float)))).number_format = '$#,##0'
    ws3.cell(row=r, column=5, value=sum(d[4] for d in opt_data if isinstance(d[4], (int, float)))).number_format = '$#,##0'
    ws3.cell(row=r, column=6, value=sum(d[3] for d in opt_data if isinstance(d[3], (int, float)))*12).number_format = '$#,##0'
    ws3.cell(row=r, column=7, value=sum(d[4] for d in opt_data if isinstance(d[4], (int, float)))*12).number_format = '$#,##0'
    ws3.cell(row=r, column=8, value="Excludes Stripe fees")
    total_row(ws3, r, 8, opt_color)
    r += 2

auto_width(ws3, 8)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: AWS SCALING 100-5000
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("AWS Scaling 100-5000")
ws4.sheet_properties.tabColor = "E65100"

r = 1
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws4.cell(row=r, column=1, value="FULL AWS: SCALING FROM 100 TO 5,000 USERS").font = title_font
ws4.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 1
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws4.cell(row=r, column=1, value="Detailed cost breakdown at each growth stage | What changes and when").font = subtitle_font
ws4.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 2

tiers = [
    ("TIER 1: 100 Users (Getting Started)", "2E7D32", [
        ["Frontend", "AWS Amplify", "SSR Hosting (Free Tier)", 0, 0, "Free: 5GB storage, 15GB transfer", "Yes - 100 users easily", "When bandwidth > 15GB/mo"],
        ["Backend", "AWS App Runner", "0.25 vCPU / 0.5 GB", 16, 192, "Smallest config, auto-scales", "Yes - ~50 concurrent req", "When CPU > 70% sustained"],
        ["Database", "AWS RDS PostgreSQL", "db.t4g.micro (FREE 12mo)", 0, 0, "Free tier for 12 months!", "Yes - 100 users, light queries", "After 12 months or DB > 20GB"],
        ["Email", "AWS SES", "Free Tier", 0, 0, "3,000 emails/mo free", "Yes - ~500 emails/mo", "When sending > 3,000/mo"],
        ["Domain & SSL", "Route 53 + ACM", "1 hosted zone", 1.50, 18, "ACM free, Route 53 ~$1.50", "Yes", "N/A"],
        ["File Storage", "AWS S3", "1 GB", 0.10, 1, "Profile photos, small uploads", "Yes - ~1GB", "When storage > 10GB"],
        ["CDN", "CloudFront", "Light traffic", 0.50, 6, "Caches static assets", "Yes", "When transfer > 1TB/mo"],
        ["Monitoring", "CloudWatch", "Basic (Free Tier)", 0, 0, "10 custom metrics, 5GB logs free", "Yes - basic monitoring", "When you need dashboards"],
        ["Secrets", "AWS Secrets Manager", "2 secrets", 0.20, 2.40, "DB password, API key", "Yes", "When secrets > 6"],
        ["Backups", "RDS Snapshots", "Automated (7 days)", 0, 0, "Free tier includes 1 snapshot", "Yes", "When you need > 7 day retention"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction", "Yes", "N/A"],
    ]),
    ("TIER 2: 500 Users (Growing)", "1565C0", [
        ["Frontend", "AWS Amplify", "SSR Hosting (Paid)", 15, 180, "Paid tier: unlimited bandwidth", "Yes - 500 users easily", "When you need custom domains"],
        ["Backend", "AWS App Runner", "0.5 vCPU / 1 GB", 25, 300, "Upgraded from 0.25 vCPU", "Yes - ~200 concurrent req", "When CPU > 70% sustained"],
        ["Database", "AWS RDS PostgreSQL", "db.t4g.micro (Paid)", 12, 144, "After free tier expires", "Yes - 500 users, moderate queries", "When DB > 20GB or CPU > 80%"],
        ["Email", "AWS SES", "Production", 2, 24, "~5,000 emails/mo at $0.10/1K", "Yes - ~2,500 emails/mo", "When sending > 10K/mo"],
        ["Domain & SSL", "Route 53 + ACM", "1 hosted zone", 1.50, 18, "Same as Tier 1", "Yes", "N/A"],
        ["File Storage", "AWS S3", "5 GB", 0.25, 3, "Profile photos, uploads", "Yes - ~5GB", "When storage > 20GB"],
        ["CDN", "CloudFront", "Moderate traffic", 2, 24, "More traffic = more CDN cost", "Yes", "When transfer > 2TB/mo"],
        ["Monitoring", "CloudWatch", "Basic + 3 Alarms", 3, 36, "Custom metrics + alarms", "Yes - basic alerting", "When you need dashboards"],
        ["Secrets", "AWS Secrets Manager", "4 secrets", 0.40, 5, "DB password, API keys, JWT, Stripe", "Yes", "When secrets > 6"],
        ["Backups", "RDS Snapshots", "7-day retention", 1, 12, "Automated daily snapshots", "Yes", "When you need 30-day retention"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction", "Yes", "N/A"],
    ]),
    ("TIER 3: 1,000 Users (Established)", "6A1B9A", [
        ["Frontend", "AWS Amplify", "SSR Hosting (Paid)", 15, 180, "Same as Tier 2", "Yes - 1,000 users", "When you need multi-region"],
        ["Backend", "AWS App Runner", "1 vCPU / 2 GB", 50, 600, "Upgraded: 2x compute + memory", "Yes - ~500 concurrent req", "When CPU > 70% or need > 2GB"],
        ["Database", "AWS RDS PostgreSQL", "db.t4g.small (2 GB)", 30, 360, "Upgraded from micro for more RAM", "Yes - 1,000 users, heavier queries", "When DB > 50GB or need more RAM"],
        ["Email", "AWS SES", "Production", 5, 60, "~10,000 emails/mo", "Yes - ~5,000 emails/mo", "When sending > 50K/mo"],
        ["Domain & SSL", "Route 53 + ACM", "1 hosted zone", 1.50, 18, "Same", "Yes", "N/A"],
        ["File Storage", "AWS S3", "10 GB", 0.50, 6, "More photos, uploads", "Yes - ~10GB", "When storage > 50GB"],
        ["CDN", "CloudFront", "Moderate-heavy", 5, 60, "More static asset caching", "Yes", "When transfer > 5TB/mo"],
        ["Monitoring", "CloudWatch + Sentry", "Standard + Dashboards", 10, 120, "Custom dashboards, more metrics", "Yes - full observability", "When you need PagerDuty"],
        ["Secrets", "AWS Secrets Manager", "6 secrets", 0.60, 7, "More API keys, tokens", "Yes", "N/A"],
        ["Backups", "RDS Snapshots", "14-day retention", 2, 24, "Longer retention for safety", "Yes", "When you need 30-day retention"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction", "Yes", "N/A"],
    ]),
    ("TIER 4: 2,500 Users (Scaling Up)", "E65100", [
        ["Frontend", "AWS Amplify", "SSR Hosting (Paid)", 15, 180, "Still handles it fine", "Yes - 2,500 users", "Consider CloudFront custom"],
        ["Backend", "AWS App Runner", "2 vCPU / 4 GB x2", 100, 1200, "2 instances for redundancy", "Yes - ~1,000 concurrent req", "When you need > 4GB or 3+ instances"],
        ["Database", "AWS RDS PostgreSQL", "db.t4g.medium (4 GB)", 60, 720, "Upgraded for more RAM + connections", "Yes - 2,500 users, complex queries", "When DB > 100GB or need read replicas"],
        ["Email", "AWS SES", "Production", 10, 120, "~25,000 emails/mo", "Yes - ~12,500 emails/mo", "When sending > 100K/mo"],
        ["Domain & SSL", "Route 53 + ACM", "1 hosted zone", 1.50, 18, "Same", "Yes", "N/A"],
        ["File Storage", "AWS S3", "25 GB", 1, 12, "More photos, videos, uploads", "Yes - ~25GB", "When storage > 100GB"],
        ["CDN", "CloudFront", "Heavy traffic", 10, 120, "Significant static asset caching", "Yes", "When transfer > 10TB/mo"],
        ["Monitoring", "CloudWatch + Sentry", "Standard + Error Tracking", 25, 300, "Full observability + error tracking", "Yes - production-grade", "When you need APM/tracing"],
        ["Secrets", "AWS Secrets Manager", "8 secrets", 0.80, 10, "More service keys", "Yes", "N/A"],
        ["Backups", "RDS + S3 Versioning", "30-day retention", 5, 60, "Extended retention + S3 versioning", "Yes", "N/A"],
        ["Redis Cache", "AWS ElastiCache", "cache.t4g.micro", 12, 144, "Session cache, API response cache", "Yes - reduces DB load 50%+", "When cache hit rate < 80%"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction", "Yes", "N/A"],
    ]),
    ("TIER 5: 5,000 Users (Enterprise Scale)", "B71C1C", [
        ["Frontend", "AWS Amplify", "SSR Hosting (Paid)", 15, 180, "Still handles it fine", "Yes - 5,000 users", "Consider CloudFront custom distro"],
        ["Backend", "AWS App Runner", "2 vCPU / 4 GB x3", 150, 1800, "3 instances for high availability", "Yes - ~2,000 concurrent req", "Consider ECS/Fargate for more control"],
        ["Database", "AWS RDS PostgreSQL", "db.r6g.large (16 GB)", 145, 1740, "Upgraded to r6g with read replica", "Yes - 5,000 users, heavy queries", "When you need multi-AZ"],
        ["Read Replica", "AWS RDS PostgreSQL", "db.t4g.medium (4 GB)", 60, 720, "Offload read queries", "Yes - read-heavy workloads", "Add more replicas as needed"],
        ["Email", "AWS SES", "Production", 20, 240, "~50,000 emails/mo", "Yes - ~25,000 emails/mo", "When sending > 100K/mo"],
        ["Domain & SSL", "Route 53 + ACM", "1 hosted zone", 1.50, 18, "Same", "Yes", "N/A"],
        ["File Storage", "AWS S3", "50 GB", 2, 24, "Photos, videos, documents", "Yes - ~50GB", "When storage > 200GB"],
        ["CDN", "CloudFront", "Heavy traffic", 20, 240, "Significant CDN traffic", "Yes", "When transfer > 20TB/mo"],
        ["Monitoring", "CloudWatch + Sentry + PagerDuty", "Full observability", 40, 480, "Dashboards, alerts, on-call rotation", "Yes - enterprise monitoring", "N/A"],
        ["Secrets", "AWS Secrets Manager", "10 secrets", 1, 12, "All service keys + rotation", "Yes", "N/A"],
        ["Backups", "RDS + S3 + Cross-Region", "35-day retention", 10, 120, "Extended retention + cross-region DR", "Yes", "N/A"],
        ["Redis Cache", "AWS ElastiCache", "cache.t4g.small", 25, 300, "Session cache, API response cache", "Yes - reduces DB load 50%+", "When cache hit rate < 80%"],
        ["Payment", "Stripe", "Pay-per-tx", 0, 0, "2.9% + $0.30 per transaction", "Yes", "N/A"],
    ]),
]

for tier_name, tier_color, tier_data in tiers:
    section_header(ws4, r, 8, tier_name, tier_color)
    r += 1
    for i, h in enumerate(["Category", "AWS Service", "Config", "Monthly", "Annual", "Why This Tier", "Can It Handle?", "Upgrade Trigger"], 1):
        ws4.cell(row=r, column=i, value=h)
    hdr(ws4, r, 8, tier_color)
    r += 1
    for i, d in enumerate(tier_data):
        for j, val in enumerate(d):
            ws4.cell(row=r, column=j+1, value=val)
            if j in [3, 4] and isinstance(val, (int, float)):
                ws4.cell(row=r, column=j+1).number_format = '$#,##0'
                ws4.cell(row=r, column=j+1).font = money_font
        row_style(ws4, r, 8, alt=(i % 2 == 1))
        r += 1
    # Total
    ws4.cell(row=r, column=1, value="TIER TOTAL")
    ws4.cell(row=r, column=4, value=sum(d[3] for d in tier_data if isinstance(d[3], (int, float)))).number_format = '$#,##0'
    ws4.cell(row=r, column=5, value=sum(d[4] for d in tier_data if isinstance(d[4], (int, float)))).number_format = '$#,##0'
    total_row(ws4, r, 8, tier_color)
    r += 2

# Summary section
section_header(ws4, r, 8, "COST PER USER SUMMARY", GREEN_DARK)
r += 1
for i, h in enumerate(["Metric", "100 Users", "500 Users", "1,000 Users", "2,500 Users", "5,000 Users", "", ""], 1):
    ws4.cell(row=r, column=i, value=h)
hdr(ws4, r, 8, GREEN_DARK)
r += 1

summary = [
    ["Total Monthly Cost", "$18", "$62", "$120", "$228", "$477", "", ""],
    ["Cost Per User/Month", "$0.18", "$0.12", "$0.12", "$0.09", "$0.10", "", ""],
    ["Annual Cost", "$216", "$744", "$1,440", "$2,736", "$5,724", "", ""],
    ["App Runner Config", "0.25 vCPU/0.5GB", "0.5 vCPU/1GB", "1 vCPU/2GB", "2 vCPU/4GB x2", "2 vCPU/4GB x3", "", ""],
    ["RDS Config", "db.t4g.micro", "db.t4g.micro", "db.t4g.small", "db.t4g.medium", "db.r6g.large+replica", "", ""],
    ["RDS RAM", "1 GB", "1 GB", "2 GB", "4 GB", "16 GB + 4 GB replica", "", ""],
    ["Max Concurrent Users", "~50", "~200", "~500", "~1,000", "~2,000", "", ""],
    ["Emails/Month (est.)", "~500", "~2,500", "~5,000", "~12,500", "~25,000", "", ""],
    ["Storage Needed", "~1 GB", "~5 GB", "~10 GB", "~25 GB", "~50 GB", "", ""],
    ["Stripe Fees (est.)", "~$30", "~$150", "~$300", "~$750", "~$1,500", "", ""],
]

for i, d in enumerate(summary):
    for j, val in enumerate(d):
        ws4.cell(row=r, column=j+1, value=val)
    row_style(ws4, r, 8, alt=(i % 2 == 1))
    ws4.cell(row=r, column=1).font = bold_font
    r += 1

r += 1
section_header(ws4, r, 8, "KEY SCALING INSIGHTS", "FF6F00")
r += 1

insights = [
    ["1. Free Tier Advantage", "First 12 months: RDS micro is FREE. App Runner has a free trial. Total cost can be as low as $18/mo for 100 users.", "", "", "", "", "", ""],
    ["2. Biggest Cost Jumps", "RDS is the biggest cost driver. Going from micro ($0 free/$12 paid) to small ($30) to medium ($60) to large ($145) is the main scaling expense.", "", "", "", "", "", ""],
    ["3. App Runner Auto-Scales", "App Runner automatically scales instances based on traffic. You only pay for what you use. At low traffic, it scales down to minimum.", "", "", "", "", "", ""],
    ["4. Stripe Scales With You", "Stripe fees are percentage-based (2.9% + $0.30/tx), so they scale automatically. At 5,000 users paying $30/mo each, that's ~$1,500/mo in Stripe fees on $50K revenue.", "", "", "", "", "", ""],
    ["5. When to Add Redis", "At 2,500+ users, add ElastiCache (Redis) to cache frequent queries. This reduces DB load by 50%+ and costs only ~$12/mo for the smallest instance.", "", "", "", "", "", ""],
    ["6. When to Add Read Replicas", "At 5,000+ users, add a read replica to offload SELECT queries from the primary DB. This costs ~$60/mo but doubles your read capacity.", "", "", "", "", "", ""],
    ["7. Multi-AZ for HA", "At 2,500+ users, consider Multi-AZ RDS (2x DB cost) for automatic failover. If the primary DB goes down, the standby takes over in < 2 minutes.", "", "", "", "", "", ""],
    ["8. Cost Per User Decreases", "AWS costs scale sub-linearly. At 100 users you pay $0.18/user/mo. At 5,000 users you pay $0.10/user/mo. The more users, the cheaper per user.", "", "", "", "", "", ""],
    ["9. 10,000+ Users", "Consider ECS/Fargate instead of App Runner for more control. Add Multi-AZ RDS. Consider Aurora Serverless for auto-scaling DB. Total: ~$800-1,200/mo.", "", "", "", "", "", ""],
]

for i, d in enumerate(insights):
    for j, val in enumerate(d):
        ws4.cell(row=r, column=j+1, value=val)
    row_style(ws4, r, 8, alt=(i % 2 == 1))
    ws4.cell(row=r, column=1).font = bold_font
    r += 1

auto_width(ws4, 8)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: RECOMMENDATION
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Recommendation")
ws5.sheet_properties.tabColor = "FF6F00"

r = 1
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws5.cell(row=r, column=1, value="RECOMMENDATION & DECISION GUIDE").font = title_font
ws5.cell(row=r, column=1).alignment = Alignment(horizontal="center")
r += 2

# Top recommendation
section_header(ws5, r, 6, "RECOMMENDED: Hybrid (Vercel + AWS Backend)", "7B1FA2")
r += 2
ws5.merge_cells(start_row=r, start_column=1, end_row=r+6, end_column=6)
ws5.cell(row=r, column=1, value=(
    "Why Hybrid?\n"
    "\n"
    "1. Vercel Pro ($20/mo) gives you unlimited bandwidth, instant deploys, preview URLs, and the best frontend DX.\n"
    "2. AWS App Runner + RDS gives you production-grade backend reliability with no cold starts.\n"
    "3. AWS SES gives you professional email delivery (3,000 free emails/mo).\n"
    "4. You only need to set up AWS for the backend - Vercel handles all frontend concerns.\n"
    "5. At ~$49-88/mo, it's cheaper than full AWS but more reliable than staying on Render free tier.\n"
    "\n"
    "Migration path: Upgrade Vercel to Pro first (5 min), then set up AWS backend (1-2 days)."
)).font = normal_font
ws5.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
r += 8

section_header(ws5, r, 6, "BUDGET PICK: Vercel + Render (Upgraded)", "00695C")
r += 2
ws5.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=6)
ws5.cell(row=r, column=1, value=(
    "If budget is the primary concern:\n"
    "\n"
    "1. Upgrade Render to Standard ($25/mo) - eliminates cold starts and sleep.\n"
    "2. Upgrade Render PostgreSQL to Starter ($7/mo) - gets you real Postgres.\n"
    "3. Add Resend for email (free tier: 3,000 emails/mo).\n"
    "4. Total: ~$32-52/mo with much better reliability than the free tier.\n"
    "\n"
    "This is the fastest path to production quality with minimal changes."
)).font = normal_font
ws5.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
r += 6

section_header(ws5, r, 6, "CHEAPEST: Self-Hosted VPS or Coolify", "BF360C")
r += 2
ws5.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=6)
ws5.cell(row=r, column=1, value=(
    "If you want the absolute lowest cost:\n"
    "\n"
    "1. Self-Hosted VPS: ~$6-10/mo for everything. You manage security, backups, and updates.\n"
    "2. Coolify: ~$10-15/mo. Same VPS but with a Heroku-like dashboard. Much easier management.\n"
    "3. Both require Linux knowledge and 1-2 days of setup.\n"
    "\n"
    "Best for: Developers comfortable with server management who want maximum control."
)).font = normal_font
ws5.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
r += 6

section_header(ws5, r, 6, "MODERN STACK: Vercel + Supabase (requires backend rewrite)", "FF6D00")
r += 2
ws5.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=6)
ws5.cell(row=r, column=1, value=(
    "If you're willing to rewrite your backend:\n"
    "\n"
    "1. Replace FastAPI with Supabase Edge Functions (TypeScript/Deno).\n"
    "2. Get auth, database, storage, and real-time all from Supabase.\n"
    "3. Vercel handles the frontend with the best DX available.\n"
    "4. Total: ~$45-75/mo with far fewer services to manage.\n"
    "\n"
    "Best for: New projects or teams willing to invest in a rewrite for long-term simplicity."
)).font = normal_font
ws5.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
r += 6

# Decision matrix
section_header(ws5, r, 6, "DECISION MATRIX: Pick Based on Your Priority", GREEN_DARK)
r += 1
for i, h in enumerate(["Your Priority", "Best Option", "Runner-Up", "Monthly Cost", "Why", ""], 1):
    ws5.cell(row=r, column=i, value=h)
hdr(ws5, r, 6)
r += 1

decisions = [
    ["Lowest Cost", "Self-Hosted VPS", "Coolify", "$6-15/mo", "VPS gives you everything for cheap, Coolify adds a GUI"],
    ["Easiest Setup", "Vercel+Render Upgraded", "Heroku", "$52-80/mo", "Just upgrade existing accounts, no migration needed"],
    ["Best Reliability", "Full AWS", "Hybrid (Vercel+AWS)", "$40-99/mo", "AWS has the best uptime, monitoring, and compliance"],
    ["Best DX (keep FastAPI)", "Vercel+Railway", "Vercel+Render Upgraded", "$30-80/mo", "Railway has better DX than Render, similar pricing"],
    ["Best DX (rewrite OK)", "Vercel+Supabase", "Netlify+Supabase", "$45-111/mo", "Supabase gives you auth, DB, storage in one platform"],
    ["Most Scalable", "Full AWS", "Google Cloud (GCP)", "$40-99/mo", "AWS/GCP have the best auto-scaling and global reach"],
    ["Fastest Global", "Cloudflare+Neon", "Fly.io", "$24-65/mo", "Cloudflare has 300+ edge locations, Fly.io has 30+"],
    ["Best for Learning", "Coolify (Self-Host)", "Self-Hosted VPS", "$10-15/mo", "Learn DevOps on your own PaaS with a safety net"],
    ["Best DB Workflow", "Vercel+PlanetScale", "Vercel+Neon DB", "$74-122/mo", "PlanetScale branching is amazing for team workflows"],
    ["Enterprise Ready", "Full AWS", "Microsoft Azure", "$40-99/mo", "Compliance, SLAs, support contracts, VPC networking"],
]

for i, d in enumerate(decisions):
    for j, val in enumerate(d):
        ws5.cell(row=r, column=j+1, value=val)
    row_style(ws5, r, 6, alt=(i % 2 == 1))
    ws5.cell(row=r, column=1).font = bold_font
    r += 1

auto_width(ws5, 6)
ws5.column_dimensions["E"].width = 55

# Save
filepath = "/Users/pranoy/Tech Projects/Tech Apps/GinasWeb/Ginas_Tennis_World_Pricing.xlsx"
wb.save(filepath)
print(f"Excel file saved to: {filepath}")
print(f"Total sheets: {len(wb.sheetnames)}")
for name in wb.sheetnames:
    print(f"  - {name}")